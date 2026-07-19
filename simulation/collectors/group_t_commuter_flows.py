"""
group_t_commuter_flows.py — Intra-country regional commuter flow matrices.

Collects within-country commuter OD data for epidemic spread comparison
across US states, DE Bundesländer, and JP prefectures.  These flows capture
how flu propagates between sub-national regions via daily worker mobility.

Sources
-------
US : US Census Bureau ACS 2016-2020 county-to-county commuting flows
     https://www2.census.gov/programs-surveys/demo/tables/metro-micro/2020/
              commuting-flows-2020/table1.xlsx
     → aggregated to state × state (51+DC postal codes; PR included)
     → origin/destination codes: 2-letter postal (matches cdc_fluview_state_activity.region)

DE : Bundesagentur für Arbeit, Pendlerverflechtungen Bundesland×Bundesland
     Reference date: 30.06.2020  (sheet "30.6.2020")
     → 16 Bundesländer × 16 Bundesländer full matrix
     → origin/destination: canonical German names (matches rki_bundesland.region)

JP : e-Stat Population Census 2020, prefecture × prefecture commuting OD
     Requires env var ESTAT_APP_ID (free at https://www.e-stat.go.jp/api/).
     Skipped (warning logged) if ESTAT_APP_ID is absent.
     → origin/destination: 2-digit prefecture code (01-47)

DB table written
----------------
commuter_flows
  (source, country, origin, destination, workers, year, collected_at)
  PK = (source, country, origin, destination)

Matrix semantics
----------------
  origin      = residence region (Wohnort / home state)
  destination = workplace region (Arbeitsort / work state)
  workers     = number of workers commuting origin → destination daily

G-116/G-117 compliant: uses safe_connect exclusively (no direct sqlite3 raw calls).
"""

from __future__ import annotations

import io
import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, TYPE_CHECKING

if TYPE_CHECKING:
    from sqlite3 import Connection as _Conn

log = logging.getLogger(__name__)

# ── constants ─────────────────────────────────────────────────────────────────

_US_EXCEL_URL = (
    "https://www2.census.gov/programs-surveys/demo/tables/metro-micro/2020"
    "/commuting-flows-2020/table1.xlsx"
)
_DE_EXCEL_URL = (
    "https://statistik.arbeitsagentur.de/Statistikdaten/Detail/202006/iiia6"
    "/beschaeftigung-pendler-blxbl/blxbl-d-0-202006-xls.xlsx"
    "?__blob=publicationFile&v=2"
)
_REQUEST_TIMEOUT_S = 120

# DE worksheet name (date of reference)
_DE_DATA_SHEET = "30.6.2020"
# 0-indexed row in openpyxl (row 12 in Excel = index 11)
_DE_DATA_ROW_START = 11
_DE_DATA_ROW_END = 26   # inclusive; index 26 = row 27 = Thüringen (last of 16)
# 0-indexed column where residence BL values begin (col D in Excel)
_DE_ORIGIN_COL_START = 3

# US state FIPS → 2-letter postal code
_FIPS_TO_POSTAL: dict[int, str] = {
    1: "AL",  2: "AK",  4: "AZ",  5: "AR",  6: "CA",  8: "CO",  9: "CT",
    10: "DE", 11: "DC", 12: "FL", 13: "GA", 15: "HI", 16: "ID", 17: "IL",
    18: "IN", 19: "IA", 20: "KS", 21: "KY", 22: "LA", 23: "ME", 24: "MD",
    25: "MA", 26: "MI", 27: "MN", 28: "MS", 29: "MO", 30: "MT", 31: "NE",
    32: "NV", 33: "NH", 34: "NJ", 35: "NM", 36: "NY", 37: "NC", 38: "ND",
    39: "OH", 40: "OK", 41: "OR", 42: "PA", 44: "RI", 45: "SC", 46: "SD",
    47: "TN", 48: "TX", 49: "UT", 50: "VT", 51: "VA", 53: "WA", 54: "WV",
    55: "WI", 56: "WY", 72: "PR",
}

# DE Bundesland 2-digit ID → canonical name (matches rki_bundesland.region)
_DE_BL_CANONICAL: dict[str, str] = {
    "01": "Schleswig-Holstein",
    "02": "Hamburg",
    "03": "Niedersachsen",
    "04": "Bremen",
    "05": "Nordrhein-Westfalen",
    "06": "Hessen",
    "07": "Rheinland-Pfalz",
    "08": "Baden-Württemberg",
    "09": "Bayern",
    "10": "Saarland",
    "11": "Berlin",
    "12": "Brandenburg",
    "13": "Mecklenburg-Vorpommern",
    "14": "Sachsen",
    "15": "Sachsen-Anhalt",
    "16": "Thüringen",
}

# Column order in DE matrix: cols 3–18 (0-indexed) correspond to BL IDs 01–16
# (residence states, same sequence as row sequence)
_DE_COL_BL_ORDER = [
    "01", "02", "03", "04", "05", "06", "07", "08",
    "09", "10", "11", "12", "13", "14", "15", "16",
]


# ── lazy DB import (G-116/G-117) ──────────────────────────────────────────────

def _safe_connect_import():
    from simulation.database import safe_connect
    return safe_connect


# ── table initialisation ──────────────────────────────────────────────────────

def _ensure_table(con: "_Conn") -> None:
    """Create commuter_flows table and index if they do not exist.

    Args:
        con: active DB connection from safe_connect.

    Side effects:
        DDL executed on con; commit issued.
    """
    con.execute("""
        CREATE TABLE IF NOT EXISTS commuter_flows (
            source      TEXT    NOT NULL,
            country     TEXT    NOT NULL,
            origin      TEXT    NOT NULL,
            destination TEXT    NOT NULL,
            workers     INTEGER NOT NULL,
            year        INTEGER NOT NULL,
            collected_at TEXT,
            PRIMARY KEY (source, country, origin, destination)
        )
    """)
    con.execute(
        "CREATE INDEX IF NOT EXISTS idx_cf_country "
        "ON commuter_flows(country, origin, destination)"
    )
    con.commit()


# ── upsert helper ─────────────────────────────────────────────────────────────

def _upsert_rows(con: "_Conn", rows: list[dict]) -> int:
    """Upsert commuter flow rows in chunks of 2 000.

    Args:
        con:  active DB connection.
        rows: list of dicts with keys matching commuter_flows schema.

    Returns:
        Total rowcount reported by executemany (may undercount on REPLACE).

    Side effects:
        Writes to commuter_flows; commits after each chunk.
    """
    if not rows:
        return 0

    sql = """
        INSERT OR REPLACE INTO commuter_flows
          (source, country, origin, destination, workers, year, collected_at)
        VALUES
          (:source, :country, :origin, :destination, :workers, :year, :collected_at)
    """
    inserted = 0
    chunk_size = 2_000
    for i in range(0, len(rows), chunk_size):
        chunk = rows[i : i + chunk_size]
        cur = con.executemany(sql, chunk)
        inserted += cur.rowcount if cur.rowcount >= 0 else len(chunk)
        con.commit()
    return inserted


# ── US sub-collector ──────────────────────────────────────────────────────────

def _fetch_us_commuter_flows() -> list[dict]:
    """Download Census ACS 2016-2020 and aggregate to state × state flows.

    Downloads table1.xlsx (county × county, ~6 MB), reads raw rows via
    openpyxl, then uses Polars groupby to aggregate to state postal-code
    pairs.  No pandas used (engineering principle #3: Polars-first).

    Excel column layout (0-indexed after 7 metadata rows skipped):
      col 0: Residence State FIPS  (str, 2-char "01"…"56")
      col 4: Workplace State FIPS  (str, 3-char "001"…"056")
      col 8: Workers in Commuting Flow (int or None for suppressed)

    FIPS → postal mapping: int("01")==int("001")==1 → both normalise to
    the same integer key in _FIPS_TO_POSTAL.

    Returns:
        List of flow dicts with keys:
          source, country, origin, destination, workers, year, collected_at.
        origin/destination are 2-letter postal codes (e.g. "CA", "NY").

    Raises:
        requests.HTTPError: if download fails.
        ImportError: if openpyxl or polars not installed.

    Performance: ~6 MB download + Polars groupby; ~15–25 s total.
    Side effects: None (pure fetch + aggregation).
    Caller responsibility: caller handles exceptions and logs.
    """
    import openpyxl
    import polars as pl
    import requests

    log.info("[T/US] downloading Census ACS commuter flows (~6 MB)…")
    resp = requests.get(_US_EXCEL_URL, timeout=_REQUEST_TIMEOUT_S)
    resp.raise_for_status()
    log.info(f"[T/US] downloaded {len(resp.content) / 1e6:.1f} MB")

    # Read raw rows with openpyxl (consistent with _fetch_de_commuter_flows)
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    for _ in range(7):        # skip 7 metadata rows
        next(rows_iter)
    next(rows_iter)            # skip header row (we use fixed col indices)

    data = list(rows_iter)

    # Build Polars DataFrame with only the 3 columns needed
    df = pl.DataFrame({
        "orig_fips": pl.Series([r[0] for r in data], dtype=pl.String),
        "dest_fips": pl.Series([r[4] for r in data], dtype=pl.String),
        "workers":   pl.Series([r[8] for r in data], dtype=pl.Int64),
    })

    # FIPS → postal-code join table (int keys, works for "01"→1 and "001"→1)
    fips_map = pl.DataFrame({
        "fips": pl.Series(list(_FIPS_TO_POSTAL.keys()), dtype=pl.Int32),
        "code": list(_FIPS_TO_POSTAL.values()),
    })

    agg = (
        df
        .with_columns([
            # cast string FIPS to int: "01"→1, "001"→1 — same integer, same key
            pl.col("orig_fips").cast(pl.Int32, strict=False),
            pl.col("dest_fips").cast(pl.Int32, strict=False),
            pl.col("workers").fill_null(0),
        ])
        .join(
            fips_map.rename({"fips": "orig_fips", "code": "origin"}),
            on="orig_fips", how="inner",
        )
        .join(
            fips_map.rename({"fips": "dest_fips", "code": "destination"}),
            on="dest_fips", how="inner",
        )
        .group_by(["origin", "destination"])
        .agg(pl.col("workers").sum())
        .filter(pl.col("workers") > 0)
    )

    now_iso = datetime.now(timezone.utc).isoformat()
    rows_out: list[dict] = [
        {
            "source":      "census_acs_2020",
            "country":     "US",
            "origin":      r["origin"],
            "destination": r["destination"],
            "workers":     r["workers"],
            "year":        2020,
            "collected_at": now_iso,
        }
        for r in agg.iter_rows(named=True)
    ]

    log.info(f"[T/US] {len(rows_out)} state×state flow pairs parsed")
    return rows_out


# ── DE sub-collector ──────────────────────────────────────────────────────────

def _fetch_de_commuter_flows() -> list[dict]:
    """Download Bundesagentur BL×BL commuter matrix and parse into long format.

    Matrix layout in sheet '30.6.2020':
      Rows 12-27 (Excel) = 16 Bundesländer as *workplace* (destination)
      Cols D-S (cols 3-18, 0-indexed) = 16 Bundesländer as *residence* (origin)
      Cell value = workers who live in col-BL and work in row-BL.

    Returns:
        List of flow dicts with keys:
          source, country, origin, destination, workers, year, collected_at.
        origin/destination are canonical German names
        (e.g. "Bayern", "Nordrhein-Westfalen") matching rki_bundesland.region.

    Raises:
        requests.HTTPError: if download fails.
        ImportError: if openpyxl not installed.
        KeyError: if sheet name not found (file changed upstream).

    Performance: ~260 KB download; 16×16 = 256 cells parsed.
    Side effects: None.
    Caller responsibility: caller handles exceptions and logs.
    """
    import openpyxl
    import requests

    log.info("[T/DE] downloading Bundesagentur BL×BL commuter matrix…")
    resp = requests.get(
        _DE_EXCEL_URL,
        timeout=_REQUEST_TIMEOUT_S,
        headers={"User-Agent": "Mozilla/5.0"},
    )
    resp.raise_for_status()
    log.info(f"[T/DE] downloaded {len(resp.content) / 1e3:.0f} KB")

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
    if _DE_DATA_SHEET not in wb.sheetnames:
        available = wb.sheetnames
        raise KeyError(
            f"[T/DE] expected sheet '{_DE_DATA_SHEET}'; found {available}"
        )
    ws = wb[_DE_DATA_SHEET]
    rows_raw = list(ws.iter_rows(values_only=True))

    now_iso = datetime.now(timezone.utc).isoformat()
    out: list[dict] = []

    for row_i in range(_DE_DATA_ROW_START, _DE_DATA_ROW_END + 1):
        row = rows_raw[row_i]
        # Col 0: BL ID (e.g. 1, "01", "1")
        raw_id = row[0]
        if raw_id is None:
            continue
        bl_id = str(int(raw_id)).zfill(2)   # "1" → "01"
        if bl_id not in _DE_BL_CANONICAL:
            continue
        dest_name = _DE_BL_CANONICAL[bl_id]  # workplace = destination

        for col_offset, origin_bl_id in enumerate(_DE_COL_BL_ORDER):
            col_idx = _DE_ORIGIN_COL_START + col_offset
            if col_idx >= len(row):
                break
            cell_val = row[col_idx]
            try:
                workers = int(cell_val)
            except (TypeError, ValueError):
                continue
            if workers <= 0:
                continue

            origin_name = _DE_BL_CANONICAL[origin_bl_id]  # residence = origin
            out.append({
                "source":      "bundesagentur_2020",
                "country":     "DE",
                "origin":      origin_name,
                "destination": dest_name,
                "workers":     workers,
                "year":        2020,
                "collected_at": now_iso,
            })

    log.info(f"[T/DE] {len(out)} BL×BL flow pairs parsed (incl. diagonal)")
    return out


# ── JP sub-collector ──────────────────────────────────────────────────────────

def _fetch_jp_commuter_flows() -> list[dict]:
    """Fetch JP prefecture × prefecture commuter flows from e-Stat 2020 Census.

    Data source: 令和2年国勢調査, statsDataId=0003454526
      "人口 従業・通学都道府県，男女，就業・通学別通勤者・通学者数－全国，都道府県（常住地）"
      (Population by work/school prefecture × gender × type, 2020 Census)

    Dimensions used:
      area    = 常住地 (residence prefecture) codes 01000–47000
      cat03   = 従業地 (work prefecture) codes 01000–47000
      cat01=0 = 総数 (total, all genders)
      cat02=11= 15歳以上通勤者 (workers aged 15+)

    Requires ESTAT_APP_ID env var (free registration at https://www.e-stat.go.jp/api/).

    Args: None.

    Returns:
        List of flow dicts with keys matching commuter_flows schema.
        origin/destination: 2-digit prefecture code strings ("01"–"47").
        Returns [] if ESTAT_APP_ID not set or API fails.

    Raises:
        Nothing — errors are logged and [] returned.

    Performance: 1 HTTPS request, ~200 KB JSON, ~5–10 s.
    Side effects: None.
    Caller responsibility: ESTAT_APP_ID env var must be set.
    """
    try:
        import requests as _req
    except ImportError:
        log.error("[T/JP] requests not installed — pip install requests")
        return []

    app_id = os.environ.get("ESTAT_APP_ID", "").strip()

    # Fallback: read from project api_key.txt (line "Japan e-stat api: <key>")
    if not app_id:
        _api_key_paths = [
            Path("simulation/data/api_key.txt"),
            Path(__file__).parents[2] / "data" / "api_key.txt",
        ]
        for _kp in _api_key_paths:
            if _kp.exists():
                for _line in _kp.read_text(encoding="utf-8").splitlines():
                    if "e-stat" in _line.lower() or "estat" in _line.lower():
                        _parts = _line.split(":", 1)
                        if len(_parts) == 2:
                            _candidate = _parts[1].strip()
                            if len(_candidate) == 40:  # e-Stat keys are 40 chars
                                app_id = _candidate
                                log.info("[T/JP] ESTAT_APP_ID loaded from api_key.txt")
                                break
                if app_id:
                    break

    if not app_id:
        log.warning(
            "[T/JP] ESTAT_APP_ID not set — JP commuter flows skipped. "
            "Register (free) at https://www.e-stat.go.jp/api/ and set "
            "ESTAT_APP_ID=<key> before re-running."
        )
        return []

    _STATS_DATA_ID = "0003454526"   # 2020 Census OD prefecture table
    _API_URL       = "https://api.e-stat.go.jp/rest/3.0/app/json/getStatsData"

    # 47 prefecture area codes (2020 census format: 01000–47000)
    _PREF_CODES = [f"{i:02d}000" for i in range(1, 48)]

    params = {
        "appId":       app_id,
        "statsDataId": _STATS_DATA_ID,
        "cdCat01":     "0",    # 総数 (all genders)
        "cdCat02":     "11",   # 15歳以上通勤者 (commuters 15+)
        "lang":        "J",
    }

    log.info("[T/JP] Fetching e-Stat 2020 census commuter OD (statsDataId=%s)", _STATS_DATA_ID)
    try:
        resp = _req.get(_API_URL, params=params, timeout=60)
        resp.raise_for_status()
    except Exception as exc:
        log.error("[T/JP] e-Stat API request failed: %s", exc)
        return []

    try:
        payload = resp.json()
    except Exception as exc:
        log.error("[T/JP] e-Stat JSON parse failed: %s", exc)
        return []

    result_code = (
        payload.get("GET_STATS_DATA", {})
               .get("RESULT", {})
               .get("STATUS", -1)
    )
    if result_code != 0:
        err_msg = (
            payload.get("GET_STATS_DATA", {})
                   .get("RESULT", {})
                   .get("ERROR_MSG", "unknown error")
        )
        log.error("[T/JP] e-Stat API returned status=%s: %s", result_code, err_msg)
        return []

    values = (
        payload.get("GET_STATS_DATA", {})
               .get("STATISTICAL_DATA", {})
               .get("DATA_INF", {})
               .get("VALUE", [])
    )
    if not isinstance(values, list):
        values = [values]

    now_iso = datetime.now(timezone.utc).isoformat()
    rows: list[dict] = []
    pref_set = set(_PREF_CODES)

    for v in values:
        origin_code = v.get("@area", "")      # residence prefecture
        dest_code   = v.get("@cat03", "")     # work prefecture
        raw_val     = v.get("$", "")

        # Skip national totals (00000) and "other" codes
        if origin_code not in pref_set or dest_code not in pref_set:
            continue

        try:
            workers = int(str(raw_val).replace(",", "").strip())
        except (ValueError, TypeError):
            continue

        if workers <= 0:
            continue

        # Convert "01000" → "01", "47000" → "47"
        origin = origin_code[:2]
        dest   = dest_code[:2]

        rows.append({
            "source":      "estat_census_2020",
            "country":     "JP",
            "origin":      origin,
            "destination": dest,
            "workers":     workers,
            "year":        2020,
            "collected_at": now_iso,
        })

    log.info(
        "[T/JP] e-Stat 2020 census: %d prefecture OD pairs fetched "
        "(%d unique origins, %d unique destinations)",
        len(rows),
        len({r["origin"] for r in rows}),
        len({r["destination"] for r in rows}),
    )
    return rows


# ── public entry point ────────────────────────────────────────────────────────

def run(
    db_path: Optional[str | Path] = None,
    skip_us: bool = False,
    skip_de: bool = False,
    skip_jp: bool = False,
) -> dict:
    """Collect intra-country commuter flow matrices and write to DB.

    Fetches US (Census ACS), DE (Bundesagentur), and JP (e-Stat) commuter OD
    data and upserts into the commuter_flows table.  Each sub-collector runs
    independently; errors are caught and reported without aborting others.

    Args:
        db_path: path to epi_real_seoul.db.  Defaults to
                 simulation/data/db/epi_real_seoul.db relative to project root.
        skip_us: skip US Census ACS collection.
        skip_de: skip DE Bundesagentur collection.
        skip_jp: skip JP e-Stat collection (also skipped if ESTAT_APP_ID unset).

    Returns:
        dict with keys:
          rows_us (int): US rows upserted.
          rows_de (int): DE rows upserted.
          rows_jp (int): JP rows upserted.
          errors  (list[str]): per-country error messages if any failed.

    Side effects:
        Writes to commuter_flows table in epi_real_seoul.db.
        Creates table and index if absent.

    Performance: ~15-30 s for US (6 MB download + Polars groupby), ~5 s for DE.
    Caller responsibility: ensure openpyxl installed (pip install openpyxl).
    """
    if db_path is None:
        from simulation.database import DB_PATH as _db_path_default
        db_path = _db_path_default

    safe_connect = _safe_connect_import()
    errors: list[str] = []
    counts: dict = {"rows_us": 0, "rows_de": 0, "rows_jp": 0, "errors": errors}

    with safe_connect(str(db_path), timeout=60.0) as con:
        _ensure_table(con)

        if not skip_us:
            try:
                us_rows = _fetch_us_commuter_flows()
                counts["rows_us"] = _upsert_rows(con, us_rows)
                log.info(f"[T/US] upserted {counts['rows_us']} rows")
            except Exception as exc:
                log.error(f"[T/US] failed: {exc}")
                errors.append(f"US: {exc}")

        if not skip_de:
            try:
                de_rows = _fetch_de_commuter_flows()
                counts["rows_de"] = _upsert_rows(con, de_rows)
                log.info(f"[T/DE] upserted {counts['rows_de']} rows")
            except Exception as exc:
                log.error(f"[T/DE] failed: {exc}")
                errors.append(f"DE: {exc}")

        if not skip_jp:
            try:
                jp_rows = _fetch_jp_commuter_flows()
                counts["rows_jp"] = _upsert_rows(con, jp_rows)
                log.info(f"[T/JP] upserted {counts['rows_jp']} rows")
            except Exception as exc:
                log.error(f"[T/JP] failed: {exc}")
                errors.append(f"JP: {exc}")

    log.info(
        f"[T] commuter_flows done — "
        f"US={counts['rows_us']} DE={counts['rows_de']} JP={counts['rows_jp']}"
        + (f" errors={errors}" if errors else "")
    )
    return counts
