"""Australia NNDSS Influenza — state/territory weekly counts.

Sprint β Item 4 full migration (Codex analysis 2026-05-26): real bodies
moved here from group_i_overseas.py. The legacy module re-exports for
back-compat.

Optional dep: `openpyxl` required for Excel parsing — falls back gracefully
when missing (returns empty + WARNING log).
"""
from __future__ import annotations

import io
import logging
from datetime import datetime, timezone

from simulation.collectors._endpoints import _NNDSS_URLS, _AIHW_CSV_URL  # noqa: F401 (AIHW future-use)
from simulation.collectors.overseas._common import (
    _connect,
    _ensure_overseas_flu_state_table,
    _resolve_db,
    _safe_int,
    _upsert_flu_state_rows,
)

log = logging.getLogger(__name__)


# AU state abbreviation → full name mapping (used for column matching)
_AU_STATES = ["NSW", "VIC", "QLD", "SA", "WA", "TAS", "NT", "ACT"]


def _parse_nndss_excel(content: bytes, now_iso: str, min_epiweek: int) -> list[dict]:
    """Parse NNDSS Excel workbook for influenza state/territory weekly counts.

    Searches all sheets for columns matching AU state abbreviations and a
    year/week indicator column.  Flexible to column ordering changes between
    annual releases.

    Args:
        content:      Raw bytes of the .xlsx file.
        now_iso:      ISO timestamp for collected_at.
        min_epiweek:  Earliest YYYYWW to include (e.g. 200801).

    Returns:
        List of row dicts for overseas_flu_state.

    Performance: O(n_sheets × n_rows) — typically <10 ms.
    Side effects: None.
    Caller responsibility: content must be valid .xlsx bytes.
    """
    try:
        import openpyxl
    except ImportError:
        log.warning("[overseas.nndss] openpyxl not installed — cannot parse NNDSS Excel")
        return []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        log.warning("[overseas.nndss] NNDSS Excel parse error: %s", e)
        return []

    rows: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        # Locate header row: first row containing ≥2 state abbreviations
        header_idx = None
        header = []
        for idx, row in enumerate(all_rows):
            row_strs = [str(c).strip().upper() if c is not None else "" for c in row]
            state_hits = sum(1 for s in _AU_STATES if s in row_strs)
            if state_hits >= 2:
                header_idx = idx
                header = row_strs
                break
        if header_idx is None:
            continue

        # Locate year/week and state columns
        week_col = next((i for i, h in enumerate(header)
                         if any(kw in h for kw in ("WEEK", "WK", "PERIOD", "DATE"))), None)
        year_col = next((i for i, h in enumerate(header)
                         if "YEAR" in h and i != week_col), None)
        state_cols = {s: header.index(s) for s in _AU_STATES if s in header}
        if not state_cols or week_col is None:
            continue

        for row in all_rows[header_idx + 1:]:
            if row[week_col] is None:
                continue
            week = _safe_int(row[week_col])
            year_v = _safe_int(row[year_col]) if year_col is not None else None
            if year_v is None:
                # Try to extract year from date-like cell
                try:
                    cell_val = row[week_col]
                    if hasattr(cell_val, "year"):
                        year_v = cell_val.year
                        week   = _safe_int(cell_val.isocalendar()[1])
                    else:
                        continue
                except Exception:
                    continue
            if year_v is None or week is None:
                continue
            epiweek = year_v * 100 + week
            if epiweek < min_epiweek:
                continue

            for state, col in state_cols.items():
                count = _safe_int(row[col])
                if count is None:
                    continue
                rows.append({
                    "country": "AU", "state": state,
                    "epiweek": epiweek,
                    "confirmed_flu_a": count,  # NNDSS does not split A/B by state
                    "confirmed_flu_b": None,
                    "total_flu":       count,
                    "population":      None,
                    "rate_per_100k":   None,
                    "collected_at":    now_iso,
                })

    log.info("[overseas.nndss] NNDSS Excel parsed: %d state-week rows", len(rows))
    return rows


def collect_au_nndss(
    db_path: str = "simulation/data/db/epi_real_seoul.db",
    start_date: str = "2008-01-01",
    end_date: str | None = None,
) -> dict:
    """Collect Australian NNDSS influenza weekly counts into overseas_flu_state.

    Tries candidate NNDSS Excel URLs.  On 403/404/timeout logs a clear WARNING
    and returns empty — does not crash.

    Args:
        db_path:    Path to epi_real_seoul.db.
        start_date: Earliest date to include (year-component used; min 2008).
        end_date:   Unused (full history per file); kept for interface uniformity.

    Returns:
        dict: rows_inserted (int), errors (list[str]), source (str, URL used).

    Performance: 1-3 HTTP requests, ~1-5 MB Excel.
    Side effects: writes overseas_flu_state.
    Caller responsibility: DB file must exist; openpyxl must be installed.
    """
    result: dict = {"rows_inserted": 0, "errors": [], "source": ""}
    try:
        start_year = int(start_date[:4])
    except (ValueError, TypeError):
        start_year = 2008
    min_epiweek = start_year * 100 + 1
    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    resolved = _resolve_db(db_path)
    con = _connect(resolved)
    try:
        _ensure_overseas_flu_state_table(con)
    except Exception as e:
        result["errors"].append(f"table setup: {e}")
        con.close()
        return result

    rows: list[dict] = []
    source_used = ""

    for url in _NNDSS_URLS:
        try:
            import requests as _req
            resp = _req.get(url, timeout=90)
            if resp.status_code in (403, 404):
                log.warning("[overseas.nndss] NNDSS: %d from %s — skipping", resp.status_code, url[:80])
                continue
            resp.raise_for_status()
            rows = _parse_nndss_excel(resp.content, now_iso, min_epiweek)
            if rows:
                source_used = url
                break
        except Exception as e:
            log.warning("[overseas.nndss] NNDSS URL %s failed: %s", url[:80], e)

    if not rows:
        msg = ("AU NNDSS: no data retrieved — all Excel URLs returned 403/404/error. "
               "Manual download from health.gov.au/nndss may be required.")
        log.warning("[overseas.nndss] %s", msg)
        result["errors"].append(msg)
        con.close()
        result["source"] = source_used
        return result

    try:
        ins, _ = _upsert_flu_state_rows(con, rows)
        result["rows_inserted"] = ins
    except Exception as e:
        result["errors"].append(f"DB upsert: {e}")
    finally:
        con.close()

    result["source"] = source_used
    log.info("[overseas.nndss] collect_au_nndss: %d inserted", result["rows_inserted"])
    return result


__all__ = [
    "_AU_STATES",
    "_parse_nndss_excel",
    "collect_au_nndss",
]
